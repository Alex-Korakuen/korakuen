#!/usr/bin/env python3
"""
Generate Excel import template files for all entity types.
Run once to create templates in imports/templates/.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# Styles
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
EXAMPLE_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
EXAMPLE_FONT = Font(italic=True, color="333333")
DESC_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
DESC_FONT = Font(size=10, color="555555")
ENUM_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
ENUM_FONT = Font(size=10, color="666666", italic=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


def create_template(filename, columns):
    """
    Create an Excel template with 4 header rows.
    columns: list of dicts with keys: name, example, description, allowed_values
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    for col_idx, col in enumerate(columns, start=1):
        # Row 1: Header
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

        # Row 2: Example data
        cell = ws.cell(row=2, column=col_idx, value=col["example"])
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER

        # Row 3: Field descriptions
        cell = ws.cell(row=3, column=col_idx, value=col["description"])
        cell.font = DESC_FONT
        cell.fill = DESC_FILL
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER

        # Row 4: Allowed values (light grey)
        allowed = col.get("allowed_values", "")
        cell = ws.cell(row=4, column=col_idx, value=allowed)
        cell.font = ENUM_FONT
        cell.fill = ENUM_FILL
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER

        # Set column width based on content
        max_len = max(
            len(str(col["name"])),
            len(str(col["example"])),
            min(len(str(col["description"])), 40),
            min(len(str(allowed)), 40) if allowed else 0,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(max_len + 4, 15)

    # Freeze top 4 rows so data entry starts below
    ws.freeze_panes = "A5"

    filepath = os.path.join(TEMPLATES_DIR, filename)
    wb.save(filepath)
    print(f"  Created: {filepath}")


# ============================================================
# Template definitions — columns match database schema exactly
# ============================================================

ENTITIES = [
    {
        "name": "entity_type",
        "example": "company",
        "description": "Required. Type of entity.",
        "allowed_values": "company | individual",
    },
    {
        "name": "document_type",
        "example": "RUC",
        "description": "Required. Type of identification document.",
        "allowed_values": "RUC | DNI | CE | Pasaporte",
    },
    {
        "name": "document_number",
        "example": "20612345678",
        "description": "Required. The actual ID number. RUC=11 digits, DNI=8 digits. Must be unique.",
        "allowed_values": "",
    },
    {
        "name": "legal_name",
        "example": "Constructora Los Andes S.A.C.",
        "description": "Required. Razón social or full legal name.",
        "allowed_values": "",
    },
    {
        "name": "common_name",
        "example": "Los Andes",
        "description": "Optional. How you refer to them day to day.",
        "allowed_values": "",
    },
    {
        "name": "city",
        "example": "Arequipa",
        "description": "Optional. City where the entity is located. Enables geographic filtering.",
        "allowed_values": "",
    },
    {
        "name": "region",
        "example": "Arequipa",
        "description": "Optional. Peruvian department/region.",
        "allowed_values": "",
    },
    {
        "name": "notes",
        "example": "Cement supplier from Arequipa",
        "description": "Optional. Free text notes.",
        "allowed_values": "",
    },
]

PROJECTS = [
    {
        "name": "project_code",
        "example": "PRY001",
        "description": "Optional. If blank, auto-generated as next sequential PRY###. If provided, must be unique and follow PRY### format.",
        "allowed_values": "PRY### format",
    },
    {
        "name": "name",
        "example": "Pavimentación Av. Los Héroes",
        "description": "Required. Project name.",
        "allowed_values": "",
    },
    {
        "name": "project_type",
        "example": "subcontractor",
        "description": "Required. Type of project.",
        "allowed_values": "subcontractor | oxi",
    },
    {
        "name": "status",
        "example": "active",
        "description": "Required. Current project status.",
        "allowed_values": "prospect | active | completed | cancelled",
    },
    {
        "name": "client_entity_document_number",
        "example": "20612345678",
        "description": "Optional. Document number of the client entity. Must exist in entities table. Nullable for prospects.",
        "allowed_values": "Lookup → entities.document_number",
    },
    {
        "name": "contract_value",
        "example": "500000.00",
        "description": "Optional. Total contract value. NUMERIC(15,2).",
        "allowed_values": "",
    },
    {
        "name": "contract_currency",
        "example": "PEN",
        "description": "Optional. Currency of contract value.",
        "allowed_values": "USD | PEN",
    },
    {
        "name": "start_date",
        "example": "2026-03-15",
        "description": "Optional. Project start date. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "expected_end_date",
        "example": "2026-09-15",
        "description": "Optional. Expected completion date. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "actual_end_date",
        "example": "",
        "description": "Optional. Actual completion date. Populated on completion. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "location",
        "example": "Lima",
        "description": "Optional. Region or city in Peru.",
        "allowed_values": "",
    },
    {
        "name": "notes",
        "example": "Subcontract under Consorcio Vial",
        "description": "Optional. Free text notes.",
        "allowed_values": "",
    },
]

QUOTES = [
    {
        "name": "project_code",
        "example": "PRY001",
        "description": "Required. Project code. Must exist in projects table.",
        "allowed_values": "Lookup → projects.project_code",
    },
    {
        "name": "entity_document_number",
        "example": "20612345678",
        "description": "Required. Document number of the quoting entity. Must exist in entities table.",
        "allowed_values": "Lookup → entities.document_number",
    },
    {
        "name": "date_received",
        "example": "2026-03-01",
        "description": "Required. Date the quote was received. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "title",
        "example": "Portland cement Type I x 42.5kg",
        "description": "Required. What was quoted.",
        "allowed_values": "",
    },
    {
        "name": "quantity",
        "example": "500",
        "description": "Optional. NUMERIC(15,4).",
        "allowed_values": "",
    },
    {
        "name": "unit_of_measure",
        "example": "bags",
        "description": "Optional. meters, units, hours, kg, bags, etc.",
        "allowed_values": "",
    },
    {
        "name": "unit_price",
        "example": "28.50",
        "description": "Optional. NUMERIC(15,4).",
        "allowed_values": "",
    },
    {
        "name": "subtotal",
        "example": "14250.00",
        "description": "Required. NUMERIC(15,2). quantity x unit_price or entered directly.",
        "allowed_values": "",
    },
    {
        "name": "igv_amount",
        "example": "2565.00",
        "description": "Optional. NUMERIC(15,2). IGV amount.",
        "allowed_values": "",
    },
    {
        "name": "total",
        "example": "16815.00",
        "description": "Required. NUMERIC(15,2). subtotal + igv_amount.",
        "allowed_values": "",
    },
    {
        "name": "currency",
        "example": "PEN",
        "description": "Required. Currency of the quote.",
        "allowed_values": "USD | PEN",
    },
    {
        "name": "exchange_rate",
        "example": "3.72",
        "description": "Required. PEN per USD rate at transaction date. NUMERIC(10,4).",
        "allowed_values": "",
    },
    {
        "name": "status",
        "example": "pending",
        "description": "Required. Quote status.",
        "allowed_values": "pending | accepted | rejected",
    },
    {
        "name": "document_ref",
        "example": "PRY001-QT-001",
        "description": "Optional. Document reference linking to SharePoint file.",
        "allowed_values": "",
    },
    {
        "name": "notes",
        "example": "Valid for 15 days",
        "description": "Optional. Reason for rejection or other context.",
        "allowed_values": "",
    },
]

COSTS = [
    # === HEADER SECTION — required fields first ===
    {
        "name": "document_ref",
        "example": "PRY001-AP-001",
        "description": "Required. Grouping key — rows with same value form one cost with multiple items.",
        "allowed_values": "",
    },
    {
        "name": "date",
        "example": "2026-03-15",
        "description": "Required. Date the expense occurred. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "title",
        "example": "Cement purchase for foundation work",
        "description": "Required. Overall invoice or expense title.",
        "allowed_values": "",
    },
    {
        "name": "bank_account",
        "example": "BCP-4567",
        "description": "Required. Bank account label. Must exist in bank_accounts table.",
        "allowed_values": "Lookup → bank_accounts.label",
    },
    {
        "name": "currency",
        "example": "PEN",
        "description": "Required. Currency of the cost.",
        "allowed_values": "USD | PEN",
    },
    {
        "name": "igv_rate",
        "example": "18",
        "description": "Required. IGV percentage. NUMERIC(5,2). Default 18.",
        "allowed_values": "Usually 18",
    },
    # === HEADER SECTION — optional fields ===
    {
        "name": "project_code",
        "example": "PRY001",
        "description": "Optional. Project code. Blank = SG&A. Derives cost_type automatically.",
        "allowed_values": "Lookup → projects.project_code",
    },
    {
        "name": "entity_document_number",
        "example": "20612345678",
        "description": "Optional. Supplier/vendor document number. Null if informal/unassigned.",
        "allowed_values": "Lookup → entities.document_number",
    },
    {
        "name": "exchange_rate",
        "example": "3.72",
        "description": "Optional. PEN per USD rate. Auto-looked up from exchange_rates table by date if blank.",
        "allowed_values": "",
    },
    {
        "name": "comprobante_type",
        "example": "factura",
        "description": "Optional. Type of payment document.",
        "allowed_values": "factura | boleta | recibo_por_honorarios | liquidacion_de_compra | planilla_jornales | none",
    },
    {
        "name": "comprobante_number",
        "example": "F001-00234",
        "description": "Optional. Payment document number.",
        "allowed_values": "",
    },
    {
        "name": "detraccion_rate",
        "example": "4",
        "description": "Optional. Detracción percentage. NUMERIC(5,2). Null if not applicable.",
        "allowed_values": "Varies by service type",
    },
    {
        "name": "payment_method",
        "example": "bank_transfer",
        "description": "Optional. How the payment was made.",
        "allowed_values": "bank_transfer | cash | check",
    },
    {
        "name": "quote_document_ref",
        "example": "PRY001-QT-001",
        "description": "Optional. Document ref of a prior accepted quote. Null if no prior quote.",
        "allowed_values": "Lookup → quotes.document_ref",
    },
    {
        "name": "due_date",
        "example": "2026-04-15",
        "description": "Optional. Payment due date. Feeds AP payment calendar. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "notes",
        "example": "Urgent delivery requested",
        "description": "Optional. Free-form context about the cost.",
        "allowed_values": "",
    },
    # === DETAIL SECTION — required fields first ===
    {
        "name": "item_title",
        "example": "Portland cement Type I x 42.5kg",
        "description": "Required. Line item title.",
        "allowed_values": "",
    },
    {
        "name": "category",
        "example": "materials",
        "description": "Required. Cost category. Different allowed values for project costs vs SG&A.",
        "allowed_values": "materials | labor | subcontractor | equipment_rental | permits_regulatory | software_licenses | partner_compensation | business_development | professional_services | office_admin | other",
    },
    {
        "name": "subtotal",
        "example": "14250.00",
        "description": "Required. NUMERIC(15,2). Line item total amount.",
        "allowed_values": "",
    },
    # === DETAIL SECTION — optional fields ===
    {
        "name": "quantity",
        "example": "500",
        "description": "Optional. NUMERIC(15,4). Null for lump sum lines.",
        "allowed_values": "",
    },
    {
        "name": "unit_of_measure",
        "example": "bags",
        "description": "Optional. meters, units, hours, kg, days, bags, etc.",
        "allowed_values": "",
    },
    {
        "name": "unit_price",
        "example": "28.50",
        "description": "Optional. NUMERIC(15,4). Null for lump sum lines.",
        "allowed_values": "",
    },
]

AR_INVOICES = [
    {
        "name": "project_code",
        "example": "PRY001",
        "description": "Required. Project code. Must exist in projects table.",
        "allowed_values": "Lookup → projects.project_code",
    },
    {
        "name": "bank_account",
        "example": "Interbank-7890",
        "description": "Required. Bank account label. Must exist in bank_accounts table.",
        "allowed_values": "Lookup → bank_accounts.label",
    },
    {
        "name": "entity_document_number",
        "example": "20198765432",
        "description": "Required. Document number of the client entity.",
        "allowed_values": "Lookup → entities.document_number",
    },
    {
        "name": "partner_company_name",
        "example": "Korakuen Ingeniería S.A.C.",
        "description": "Required. Name of the partner company that issued the invoice.",
        "allowed_values": "Lookup → partner_companies.name",
    },
    {
        "name": "invoice_number",
        "example": "F001-00089",
        "description": "Required. Own invoice numbering from Alegra/Contasis.",
        "allowed_values": "",
    },
    {
        "name": "comprobante_type",
        "example": "factura",
        "description": "Required. Type of payment document. Always factura for construction AR.",
        "allowed_values": "factura | boleta | recibo_por_honorarios",
    },
    {
        "name": "invoice_date",
        "example": "2026-04-01",
        "description": "Required. Invoice issue date. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "due_date",
        "example": "2026-05-01",
        "description": "Optional. Payment due date. Format: YYYY-MM-DD.",
        "allowed_values": "",
    },
    {
        "name": "subtotal",
        "example": "150000.00",
        "description": "Required. NUMERIC(15,2). Invoice subtotal amount.",
        "allowed_values": "",
    },
    {
        "name": "igv_rate",
        "example": "18",
        "description": "Required. IGV percentage. NUMERIC(5,2). Default 18.",
        "allowed_values": "Usually 18",
    },
    {
        "name": "detraccion_rate",
        "example": "4",
        "description": "Optional. Detracción percentage. NUMERIC(5,2). Null if not applicable.",
        "allowed_values": "Varies by service type",
    },
    {
        "name": "retencion_applicable",
        "example": "false",
        "description": "Required. Whether client will withhold retención. BOOLEAN.",
        "allowed_values": "true | false",
    },
    {
        "name": "retencion_rate",
        "example": "3",
        "description": "Optional. Retención percentage. NUMERIC(5,2). Required if retencion_applicable is true. Default 3.",
        "allowed_values": "Usually 3 (if applicable)",
    },
    {
        "name": "currency",
        "example": "PEN",
        "description": "Required. Currency of the invoice.",
        "allowed_values": "USD | PEN",
    },
    {
        "name": "exchange_rate",
        "example": "3.72",
        "description": "Required. PEN per USD rate at transaction date. NUMERIC(10,4).",
        "allowed_values": "",
    },
    {
        "name": "document_ref",
        "example": "PRY001-AR-001",
        "description": "Optional. Document reference linking to SharePoint file.",
        "allowed_values": "",
    },
    {
        "name": "retencion_verified",
        "example": "false",
        "description": "Required. Manually set to true once confirmed that client paid retención to SUNAT. BOOLEAN. Default false.",
        "allowed_values": "true | false",
    },
    {
        "name": "notes",
        "example": "March 2026 billing",
        "description": "Optional. Free text notes.",
        "allowed_values": "",
    },
]


def main():
    print("Generating import templates...\n")

    templates = [
        ("entities.xlsx", ENTITIES),
        ("projects.xlsx", PROJECTS),
        ("quotes.xlsx", QUOTES),
        ("costs.xlsx", COSTS),
        ("ar_invoices.xlsx", AR_INVOICES),
    ]

    for filename, columns in templates:
        create_template(filename, columns)

    print(f"\nDone. {len(templates)} templates created in {TEMPLATES_DIR}/")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Module: projects.py
Purpose: All project operations — add single, import from Excel, set budget
Tables: projects, project_budgets
"""

import pandas as pd
from openpyxl import load_workbook

from lib.db import supabase
from lib.helpers import (
    get_input, get_optional_input, confirm, list_choices, clear_screen,
    get_currency,
)
from lib.import_helpers import (
    DATA_START_ROW,
    clear_highlighting, apply_error_highlighting,
    validate_required, validate_enum, validate_lookup,
    validate_date, validate_number,
    print_errors,
)


def menu():
    """Submenu for project operations. Called by main.py."""
    while True:
        clear_screen()
        print("\n=== Projects ===\n")
        print("1. Add project")
        print("2. Import projects from Excel")
        print("3. Set project budget")
        print("4. Back")

        choice = get_input("\nSelect option: ")

        if choice == "1":
            add_project()
        elif choice == "2":
            import_projects()
        elif choice == "3":
            set_project_budget()
        elif choice == "4":
            return


# ============================================================
# Add Project
# ============================================================

def _next_project_code():
    """Generate the next sequential project code (PRY001, PRY002...)."""
    result = (
        supabase.table("projects")
        .select("project_code")
        .order("project_code", desc=True)
        .limit(1)
        .execute()
    )
    if result.data:
        last_code = result.data[0]["project_code"]
        last_num = int(last_code.replace("PRY", ""))
        return f"PRY{last_num + 1:03d}"
    return "PRY001"


def add_project():
    """Register a single project interactively."""
    clear_screen()
    print("\n=== Add Project ===\n")

    project_code = _next_project_code()
    print(f"  Project code: {project_code} (auto-generated)\n")

    # --- Name ---
    name = get_input("  Project name: ")

    # --- Project type ---
    print("\n  Project types: subcontractor, oxi")
    project_type = get_input("  Project type: ").lower()
    while project_type not in ("subcontractor", "oxi"):
        print("  Must be 'subcontractor' or 'oxi'.")
        project_type = get_input("  Project type: ").lower()

    # --- Status ---
    print("\n  Statuses: prospect, active, completed, cancelled")
    status = get_input("  Status: ").lower()
    while status not in ("prospect", "active", "completed", "cancelled"):
        print("  Must be prospect, active, completed, or cancelled.")
        status = get_input("  Status: ").lower()

    # --- Client entity (optional) ---
    client_entity_id = None
    client_name = None
    if confirm("\n  Assign a client entity?"):
        from modules.entities import _search_and_select_entity
        entity = _search_and_select_entity()
        if entity:
            client_entity_id = entity["id"]
            client_name = entity["legal_name"]

    # --- Contract value (optional) ---
    contract_value = get_optional_input("\n  Contract value (optional — press Enter to skip): ")
    contract_currency = None
    if contract_value:
        try:
            contract_value = float(contract_value)
        except ValueError:
            print("  Invalid number, skipping contract value.")
            contract_value = None
    if contract_value:
        print("  Currencies: USD, PEN")
        contract_currency = get_currency(label="Contract currency")

    # --- Optional fields ---
    start_date = get_optional_input("  Start date (YYYY-MM-DD, optional — press Enter to skip): ")
    location = get_optional_input("  Location (optional — press Enter to skip): ")
    notes = get_optional_input("  Notes (optional — press Enter to skip): ")

    # --- Summary ---
    print("\n--- Summary ---")
    print(f"  Code:     {project_code}")
    print(f"  Name:     {name}")
    print(f"  Type:     {project_type}")
    print(f"  Status:   {status}")
    if client_name:
        print(f"  Client:   {client_name}")
    if contract_value:
        print(f"  Contract: {contract_currency} {contract_value:,.2f}")
    if start_date:
        print(f"  Start:    {start_date}")
    if location:
        print(f"  Location: {location}")
    if notes:
        print(f"  Notes:    {notes}")

    if not confirm("\nRegister this project?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # --- Insert ---
    data = {
        "project_code": project_code,
        "name": name,
        "project_type": project_type,
        "status": status,
    }
    if client_entity_id:
        data["client_entity_id"] = client_entity_id
    if contract_value:
        data["contract_value"] = contract_value
        data["contract_currency"] = contract_currency
    if start_date:
        data["start_date"] = start_date
    if location:
        data["location"] = location
    if notes:
        data["notes"] = notes

    try:
        response = supabase.table("projects").insert(data).execute()
        print(f"\n✓ Project registered: {project_code} (ID: {response.data[0]['id'][:8]}...)")
    except Exception as e:
        print(f"\n✗ Error: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Import Projects
# ============================================================

def _load_project_lookups():
    """Pre-load lookup tables for project import validation."""
    entities = supabase.table("entities").select("id, document_number").eq("is_active", True).execute()
    existing = supabase.table("projects").select("project_code").execute()
    return {
        "entities": {r["document_number"]: r["id"] for r in entities.data},
        "existing_codes": {r["project_code"] for r in existing.data},
    }


def _validate_project_row(row_num, row, errors, lookups):
    """Validate a single project row."""
    validate_required(row_num, row, "name", errors)
    validate_required(row_num, row, "project_type", errors)
    validate_required(row_num, row, "status", errors)

    validate_enum(row_num, row, "project_type", ["subcontractor", "oxi"], errors)
    validate_enum(row_num, row, "status", ["prospect", "active", "completed", "cancelled"], errors)
    validate_enum(row_num, row, "contract_currency", ["USD", "PEN"], errors)

    validate_lookup(row_num, row, "client_entity_document_number", lookups["entities"], errors)

    validate_date(row_num, row, "start_date", errors)
    validate_date(row_num, row, "expected_end_date", errors)
    validate_date(row_num, row, "actual_end_date", errors)
    validate_number(row_num, row, "contract_value", errors)

    # Validate project_code format and uniqueness if provided
    code = row.get("project_code")
    if code and not pd.isna(code):
        code_str = str(code).strip()
        if not code_str.startswith("PRY") or len(code_str) != 6:
            errors.append((row_num, "project_code", "Must follow PRY### format"))
        elif code_str in lookups["existing_codes"]:
            errors.append((row_num, "project_code", "Already exists in database"))


def _build_project_record(row, lookups, next_code_num):
    """Convert a spreadsheet row to a database record."""
    # Auto-generate project_code if blank
    code = row.get("project_code")
    if code and not pd.isna(code):
        project_code = str(code).strip()
    else:
        project_code = f"PRY{next_code_num[0]:03d}"
        next_code_num[0] += 1

    data = {
        "project_code": project_code,
        "name": str(row["name"]).strip(),
        "project_type": str(row["project_type"]).strip(),
        "status": str(row["status"]).strip(),
    }

    # FK lookups
    client_doc = row.get("client_entity_document_number")
    if client_doc and not pd.isna(client_doc) and str(client_doc).strip():
        data["client_entity_id"] = lookups["entities"][str(client_doc).strip()]

    # Optional fields
    for field in ("contract_value",):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = float(val)

    for field in ("contract_currency",):
        val = row.get(field)
        if val is not None and not pd.isna(val) and str(val).strip():
            data[field] = str(val).strip()

    for field in ("start_date", "expected_end_date", "actual_end_date"):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = pd.Timestamp(val).strftime("%Y-%m-%d")

    for field in ("location", "notes"):
        val = row.get(field)
        if val is not None and not pd.isna(val) and str(val).strip():
            data[field] = str(val).strip()

    return data


def import_projects():
    """Import projects from an Excel spreadsheet."""
    clear_screen()
    print("\n=== Import Projects ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return

    print(f"Found {len(df)} data rows.")

    lookups = _load_project_lookups()

    # Validate all rows
    errors = []
    for idx, row in df.iterrows():
        excel_row = idx + DATA_START_ROW
        _validate_project_row(excel_row, row, errors, lookups)

    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return

    # Clear previous highlighting
    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)

    # Show summary
    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        code = row.get("project_code", "auto")
        if pd.isna(code):
            code = "auto"
        print(f"    {i+1}. {code} — {row.get('name', '')}")

    if not confirm(f"\nImport {len(df)} projects?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # Determine next code number for auto-generation
    all_codes = lookups["existing_codes"]
    if all_codes:
        max_num = max(int(c.replace("PRY", "")) for c in all_codes)
    else:
        max_num = 0
    next_code_num = [max_num + 1]  # mutable counter

    try:
        records = [_build_project_record(row, lookups, next_code_num) for _, row in df.iterrows()]
        response = supabase.table("projects").insert(records).execute()
        print(f"\n✓ {len(response.data)} projects imported successfully.")
    except Exception as e:
        print(f"\n✗ Error during import: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Set Project Budget
# ============================================================

# Must match cost_items project cost categories exactly
PROJECT_BUDGET_CATEGORIES = [
    "materials",
    "labor",
    "subcontractor",
    "equipment_rental",
    "permits_regulatory",
    "other",
]


def set_project_budget():
    """Set budget targets per category for a project."""
    clear_screen()
    print("\n=== Set Project Budget ===\n")

    # --- Select project ---
    projects = (
        supabase.table("projects")
        .select("id, project_code, name")
        .eq("is_active", True)
        .order("project_code")
        .execute()
    )
    list_choices("Active projects", projects.data, display=["project_code", "name"])

    if not projects.data:
        input("\nPress Enter to continue...")
        return

    project_code_input = get_input("Enter project code (e.g. PRY001): ").upper()
    project = next(
        (p for p in projects.data if p["project_code"] == project_code_input), None
    )
    if not project:
        print(f"\n✗ Project '{project_code_input}' not found.")
        input("\nPress Enter to continue...")
        return

    project_id = project["id"]
    project_code = project["project_code"]
    project_name = project["name"]

    # --- Check for existing budgets ---
    existing = (
        supabase.table("project_budgets")
        .select("category, budgeted_amount, currency, notes")
        .eq("project_id", project_id)
        .execute()
    )

    if existing.data:
        print(f"\n  Existing budgets for {project_code} — {project_name}:")
        for entry in existing.data:
            notes_str = f" ({entry['notes']})" if entry.get("notes") else ""
            print(f"    {entry['category']:25s} {entry['currency']} {entry['budgeted_amount']:>12,.2f}{notes_str}")
        print()
        if not confirm("  Overwrite existing budgets?"):
            print("  Cancelled.")
            input("\nPress Enter to continue...")
            return
        # Delete existing budgets before re-inserting
        try:
            supabase.table("project_budgets").delete().eq("project_id", project_id).execute()
        except Exception as e:
            print(f"\n✗ Error clearing existing budgets: {e}")
            input("\nPress Enter to continue...")
            return

    # --- Currency ---
    print("  Currencies: USD, PEN")
    currency = get_currency(label="Budget currency")

    # --- Collect amounts per category ---
    print(f"\n  Enter budgeted amount for each category ({currency}):\n")
    budget_entries = []
    for category in PROJECT_BUDGET_CATEGORIES:
        display_name = category.replace("_", " ").title()
        while True:
            amount_str = get_input(f"    {display_name}: ")
            try:
                amount = float(amount_str)
                if amount < 0:
                    print("    Amount must be zero or positive.")
                    continue
                break
            except ValueError:
                print("    Invalid number. Enter a valid amount.")
        budget_entries.append({"category": category, "amount": amount})

    # --- Notes ---
    notes = get_optional_input("\n  Notes (optional — press Enter to skip): ")

    # --- Summary ---
    print("\n--- Summary ---")
    print(f"  Project:  {project_code} — {project_name}")
    print(f"  Currency: {currency}")
    print()
    total_budget = 0
    for entry in budget_entries:
        display_name = entry["category"].replace("_", " ").title()
        print(f"    {display_name:25s} {currency} {entry['amount']:>12,.2f}")
        total_budget += entry["amount"]
    print(f"    {'':25s} {'─' * 20}")
    print(f"    {'Total':25s} {currency} {total_budget:>12,.2f}")
    if notes:
        print(f"\n  Notes: {notes}")

    if not confirm("\nSet this budget?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # --- Batch insert ---
    records = []
    for entry in budget_entries:
        data = {
            "project_id": project_id,
            "category": entry["category"],
            "budgeted_amount": entry["amount"],
            "currency": currency,
        }
        if notes:
            data["notes"] = notes
        records.append(data)

    try:
        response = supabase.table("project_budgets").insert(records).execute()
        print(f"\n✓ Budget set for {project_code}: {len(response.data)} categories ({currency} {total_budget:,.2f} total)")
    except Exception as e:
        print(f"\n✗ Error: {e}")

    input("\nPress Enter to continue...")

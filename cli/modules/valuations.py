#!/usr/bin/env python3
"""
Module: valuations.py
Purpose: All valuation operations — add single, import from Excel
Tables: valuations
"""

import pandas as pd
from openpyxl import load_workbook

from lib.db import supabase
from lib.helpers import get_input, get_optional_input, confirm, list_choices, clear_screen
from lib.import_helpers import (
    RED_FILL, NO_FILL, DATA_START_ROW,
    clear_highlighting, apply_error_highlighting,
    validate_required, validate_enum, validate_lookup,
    validate_date, validate_number,
    print_errors,
)


def menu():
    """Submenu for valuation operations. Called by main.py."""
    while True:
        clear_screen()
        print("\n=== Valuations ===\n")
        print("1. Add valuation")
        print("2. Import valuations from Excel")
        print("3. Back")

        choice = get_input("\nSelect option: ")

        if choice == "1":
            add_valuation()
        elif choice == "2":
            import_valuations()
        elif choice == "3":
            return


# ============================================================
# Add Valuation
# ============================================================

def add_valuation():
    """Register a single valuation interactively."""
    clear_screen()
    print("\n=== Add Valuation ===\n")

    # --- Select project ---
    projects = (
        supabase.table("projects")
        .select("id, project_code, name")
        .eq("is_active", True)
        .order("project_code")
        .execute()
    )
    if not projects.data:
        print("No active projects found.")
        input("\nPress Enter to continue...")
        return

    list_choices("Active projects", projects.data, display=["project_code", "name"])
    proj_num = get_input("  Select project number: ")
    try:
        project = projects.data[int(proj_num) - 1]
    except (ValueError, IndexError):
        print("\n✗ Invalid selection.")
        input("\nPress Enter to continue...")
        return

    # --- Auto-generate valuation number ---
    existing = (
        supabase.table("valuations")
        .select("valuation_number")
        .eq("project_id", project["id"])
        .order("valuation_number", desc=True)
        .limit(1)
        .execute()
    )
    if existing.data:
        valuation_number = existing.data[0]["valuation_number"] + 1
    else:
        valuation_number = 1

    print(f"\n  Valuation number: {valuation_number} (auto-generated for {project['project_code']})")

    # --- Period ---
    period_month = get_input("  Period month (1-12): ")
    while not period_month.isdigit() or not (1 <= int(period_month) <= 12):
        print("  Must be a number from 1 to 12.")
        period_month = get_input("  Period month (1-12): ")
    period_month = int(period_month)

    period_year = get_input("  Period year (e.g. 2026): ")
    while not period_year.isdigit() or len(period_year) != 4:
        print("  Must be a 4-digit year.")
        period_year = get_input("  Period year (e.g. 2026): ")
    period_year = int(period_year)

    # --- Status ---
    print("\n  Statuses: open, closed")
    status = get_input("  Status (default: open): ").lower() or "open"
    while status not in ("open", "closed"):
        print("  Must be 'open' or 'closed'.")
        status = get_input("  Status: ").lower()

    # --- Optional fields ---
    billed_value = get_optional_input("  Billed value (optional — press Enter to skip): ")
    billed_currency = None
    if billed_value:
        try:
            billed_value = float(billed_value)
        except ValueError:
            print("  Invalid number, skipping billed value.")
            billed_value = None
    if billed_value:
        print("  Currencies: USD, PEN")
        billed_currency = get_input("  Billed currency: ").upper()
        while billed_currency not in ("USD", "PEN"):
            print("  Must be USD or PEN.")
            billed_currency = get_input("  Billed currency: ").upper()

    date_closed = None
    if status == "closed":
        date_closed = get_optional_input("  Date closed (YYYY-MM-DD, optional — press Enter to skip): ")

    notes = get_optional_input("  Notes (optional — press Enter to skip): ")

    # --- Summary ---
    print("\n--- Summary ---")
    print(f"  Project:    {project['project_code']} — {project['name']}")
    print(f"  Valuation:  #{valuation_number}")
    print(f"  Period:     {period_month}/{period_year}")
    print(f"  Status:     {status}")
    if billed_value:
        print(f"  Billed:     {billed_currency} {billed_value:,.2f}")
    if date_closed:
        print(f"  Closed:     {date_closed}")
    if notes:
        print(f"  Notes:      {notes}")

    if not confirm("\nRegister this valuation?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # --- Insert ---
    data = {
        "project_id": project["id"],
        "valuation_number": valuation_number,
        "period_month": period_month,
        "period_year": period_year,
        "status": status,
    }
    if billed_value:
        data["billed_value"] = billed_value
        data["billed_currency"] = billed_currency
    if date_closed:
        data["date_closed"] = date_closed
    if notes:
        data["notes"] = notes

    try:
        response = supabase.table("valuations").insert(data).execute()
        print(f"\n✓ Valuation #{valuation_number} registered (ID: {response.data[0]['id'][:8]}...)")
    except Exception as e:
        print(f"\n✗ Error: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Import Valuations
# ============================================================

def _load_valuation_lookups():
    """Pre-load lookup tables for valuation import."""
    projects = supabase.table("projects").select("id, project_code").eq("is_active", True).execute()
    valuations = supabase.table("valuations").select("project_id, valuation_number").execute()
    return {
        "projects": {r["project_code"]: r["id"] for r in projects.data},
        "existing": {(r["project_id"], r["valuation_number"]) for r in valuations.data},
    }


def _validate_valuation_row(row_num, row, errors, lookups):
    """Validate a single valuation row."""
    validate_required(row_num, row, "project_code", errors)
    validate_required(row_num, row, "valuation_number", errors)
    validate_required(row_num, row, "period_month", errors)
    validate_required(row_num, row, "period_year", errors)
    validate_required(row_num, row, "status", errors)

    validate_enum(row_num, row, "status", ["open", "closed"], errors)
    validate_enum(row_num, row, "billed_currency", ["USD", "PEN"], errors)
    validate_lookup(row_num, row, "project_code", lookups["projects"], errors)

    validate_number(row_num, row, "valuation_number", errors)
    validate_number(row_num, row, "period_month", errors)
    validate_number(row_num, row, "period_year", errors)
    validate_number(row_num, row, "billed_value", errors)
    validate_date(row_num, row, "date_closed", errors)

    # Validate period_month range
    month = row.get("period_month")
    if month is not None and not pd.isna(month):
        try:
            m = int(float(month))
            if not (1 <= m <= 12):
                errors.append((row_num, "period_month", "Must be between 1 and 12"))
        except (ValueError, TypeError):
            pass  # already caught by validate_number

    # Check uniqueness (project_id + valuation_number)
    code = row.get("project_code")
    val_num = row.get("valuation_number")
    if code and not pd.isna(code) and val_num is not None and not pd.isna(val_num):
        code_str = str(code).strip()
        project_id = lookups["projects"].get(code_str)
        if project_id:
            try:
                vn = int(float(val_num))
                if (project_id, vn) in lookups["existing"]:
                    errors.append((row_num, "valuation_number", f"Valuation #{vn} already exists for {code_str}"))
            except (ValueError, TypeError):
                pass


def _build_valuation_record(row, lookups):
    """Convert a spreadsheet row to a database record."""
    project_id = lookups["projects"][str(row["project_code"]).strip()]

    data = {
        "project_id": project_id,
        "valuation_number": int(float(row["valuation_number"])),
        "period_month": int(float(row["period_month"])),
        "period_year": int(float(row["period_year"])),
        "status": str(row["status"]).strip(),
    }

    if not pd.isna(row.get("billed_value")) and str(row.get("billed_value", "")).strip():
        data["billed_value"] = float(row["billed_value"])
    if not pd.isna(row.get("billed_currency")) and str(row.get("billed_currency", "")).strip():
        data["billed_currency"] = str(row["billed_currency"]).strip()
    if not pd.isna(row.get("date_closed")) and str(row.get("date_closed", "")).strip():
        data["date_closed"] = pd.Timestamp(row["date_closed"]).strftime("%Y-%m-%d")
    if not pd.isna(row.get("notes")) and str(row.get("notes", "")).strip():
        data["notes"] = str(row["notes"]).strip()

    return data


def import_valuations():
    """Import valuations from an Excel spreadsheet."""
    clear_screen()
    print("\n=== Import Valuations ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return

    print(f"Found {len(df)} data rows.")

    lookups = _load_valuation_lookups()

    errors = []
    for idx, row in df.iterrows():
        excel_row = idx + DATA_START_ROW
        _validate_valuation_row(excel_row, row, errors, lookups)

    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)

    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        print(f"    {i+1}. {row.get('project_code', '')} Val #{row.get('valuation_number', '')} — {row.get('period_month', '')}/{row.get('period_year', '')}")

    if not confirm(f"\nImport {len(df)} valuations?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    try:
        records = [_build_valuation_record(row, lookups) for _, row in df.iterrows()]
        response = supabase.table("valuations").insert(records).execute()
        print(f"\n✓ {len(response.data)} valuations imported successfully.")
    except Exception as e:
        print(f"\n✗ Error during import: {e}")

    input("\nPress Enter to continue...")

#!/usr/bin/env python3
"""
Module: costs.py
Purpose: All cost operations — add single (header + line items), import from Excel
Tables: costs, cost_items
"""

import pandas as pd
from openpyxl import load_workbook

from lib.db import supabase
from lib.helpers import (
    get_input, get_optional_input, confirm, list_choices, clear_screen,
    get_currency, get_exchange_rate, select_project,
)
from lib.import_helpers import (
    DATA_START_ROW,
    clear_highlighting, apply_error_highlighting,
    validate_required, validate_enum, validate_lookup,
    validate_date, validate_number,
    print_errors,
)

# Cost item categories by cost type
PROJECT_CATEGORIES = ["materials", "labor", "subcontractor", "equipment_rental", "permits_regulatory", "other"]
SGA_CATEGORIES = ["software_licenses", "partner_compensation", "business_development", "professional_services", "office_admin", "other"]
ALL_CATEGORIES = list(set(PROJECT_CATEGORIES + SGA_CATEGORIES))


def menu():
    """Submenu for cost operations. Called by main.py."""
    while True:
        clear_screen()
        print("\n=== Costs ===\n")
        print("1. Add cost")
        print("2. Import costs from Excel")
        print("3. Import cost items from Excel")
        print("4. Back")

        choice = get_input("\nSelect option: ")

        if choice == "1":
            add_cost()
        elif choice == "2":
            import_costs()
        elif choice == "3":
            import_cost_items()
        elif choice == "4":
            return


# ============================================================
# Add Cost — Two-step: Header + Line Items
# ============================================================

def add_cost():
    """Register a single cost with line items interactively."""
    clear_screen()
    print("\n=== Add Cost ===\n")

    # ---- STEP 1: HEADER ----

    # --- Cost type ---
    print("  Cost types: project_cost, sga")
    cost_type = get_input("  Cost type: ").lower()
    while cost_type not in ("project_cost", "sga"):
        print("  Must be 'project_cost' or 'sga'.")
        cost_type = get_input("  Cost type: ").lower()

    # --- Project (optional for SGA, expected for project_cost) ---
    project = None
    if cost_type == "project_cost":
        project = select_project()
        if not project:
            return

    # --- Valuation (optional) ---
    valuation = None
    if project:
        vals = (
            supabase.table("valuations")
            .select("id, valuation_number, period_month, period_year")
            .eq("project_id", project["id"])
            .eq("status", "open")
            .order("valuation_number")
            .execute()
        )
        if vals.data:
            list_choices("Open valuations", vals.data, display=["valuation_number", "period_month", "period_year"])
            val_num = get_optional_input("  Select valuation number (optional — press Enter to skip): ")
            if val_num:
                try:
                    valuation = vals.data[int(val_num) - 1]
                except (ValueError, IndexError):
                    print("  Invalid selection, skipping valuation.")

    # --- Bank account (required) ---
    bank_accounts = (
        supabase.table("bank_accounts")
        .select("id, bank_name, account_number_last4, currency, partner_companies(name)")
        .eq("is_active", True)
        .execute()
    )
    if not bank_accounts.data:
        print("\n  No bank accounts found.")
        input("\nPress Enter to continue...")
        return

    print("\n  Available bank accounts:")
    for i, ba in enumerate(bank_accounts.data, start=1):
        partner = ba.get("partner_companies", {}).get("name", "Unknown")
        print(f"    {i}. {ba['bank_name']} {ba['currency']} {ba['account_number_last4']} ({partner})")
    print()

    bank_num = get_input("  Select bank account number: ")
    try:
        bank_account = bank_accounts.data[int(bank_num) - 1]
    except (ValueError, IndexError):
        print("\n  ✗ Invalid selection.")
        input("\nPress Enter to continue...")
        return

    # --- Entity (optional) ---
    entity = None
    if confirm("\n  Assign an entity (supplier)?"):
        from modules.entities import _search_and_select_entity
        entity = _search_and_select_entity()

    # --- Quote (optional) ---
    quote = None
    if project:
        quotes = (
            supabase.table("quotes")
            .select("id, title, document_ref, total, currency")
            .eq("project_id", project["id"])
            .eq("status", "accepted")
            .execute()
        )
        if quotes.data:
            list_choices("Accepted quotes", quotes.data, display=["document_ref", "title"])
            q_num = get_optional_input("  Select quote number (optional — press Enter to skip): ")
            if q_num:
                try:
                    quote = quotes.data[int(q_num) - 1]
                except (ValueError, IndexError):
                    print("  Invalid selection, skipping quote.")

    # --- Date, title ---
    date_str = get_input("\n  Date (YYYY-MM-DD): ")
    title = get_input("  Title: ")

    # --- Tax rates ---
    igv_input = get_input("  IGV rate % (default: 18): ") or "18"
    try:
        igv_rate = float(igv_input)
    except ValueError:
        igv_rate = 18.0

    detraccion_input = get_optional_input("  Detraccion rate % (optional — press Enter to skip): ")
    detraccion_rate = None
    if detraccion_input:
        try:
            detraccion_rate = float(detraccion_input)
        except ValueError:
            detraccion_rate = None

    # --- Currency ---
    print("\n  Currencies: USD, PEN")
    currency = get_currency()
    exchange_rate = get_exchange_rate()

    # --- Comprobante (optional) ---
    comprobante_type = None
    comprobante_number = None
    valid_comprobante_types = ("factura", "boleta", "recibo_por_honorarios", "liquidacion_de_compra", "planilla_jornales", "none")
    comp = get_optional_input("  Comprobante type (factura/boleta/recibo_por_honorarios/liquidacion_de_compra/planilla_jornales/none, optional — press Enter to skip): ")
    if comp:
        comp = comp.lower()
        if comp in valid_comprobante_types:
            comprobante_type = comp
            comprobante_number = get_optional_input("  Comprobante number (optional — press Enter to skip): ")

    # --- Payment method (optional) ---
    payment_method = None
    pm = get_optional_input("  Payment method (bank_transfer/cash/check, optional — press Enter to skip): ")
    if pm:
        pm = pm.lower()
        if pm in ("bank_transfer", "cash", "check"):
            payment_method = pm

    document_ref = get_optional_input("  Document ref (e.g. PRY001-AP-001, optional — press Enter to skip): ")
    due_date = get_optional_input("  Due date (YYYY-MM-DD, optional — press Enter to skip): ")
    notes = get_optional_input("  Notes (optional — press Enter to skip): ")

    # ---- STEP 2: LINE ITEMS ----
    print("\n--- Add Line Items ---")
    print("(Add at least one item)\n")

    categories = PROJECT_CATEGORIES if cost_type == "project_cost" else SGA_CATEGORIES
    items = []
    running_total = 0.0
    item_num = 1

    while True:
        print(f"  Item {item_num}:")

        item_title = get_input("    Title: ")

        print(f"    Categories: {', '.join(categories)}")
        category = get_input("    Category: ").lower()
        while category not in categories:
            print(f"    Must be one of: {', '.join(categories)}")
            category = get_input("    Category: ").lower()

        qty = get_optional_input("    Quantity (optional — press Enter for lump sum): ")
        uom = None
        up = None
        if qty:
            try:
                qty = float(qty)
            except ValueError:
                qty = None
        if qty:
            uom = get_optional_input("    Unit of measure (optional — press Enter to skip): ")
            up_input = get_optional_input("    Unit price (optional — press Enter to skip): ")
            if up_input:
                try:
                    up = float(up_input)
                except ValueError:
                    up = None

        # Subtotal
        default_sub = qty * up if (qty and up) else None
        if default_sub:
            print(f"    Computed: {default_sub:,.2f}")
        sub_input = get_input(f"    Subtotal{f' (default: {default_sub:,.2f})' if default_sub else ''}: ")
        try:
            item_subtotal = float(sub_input) if sub_input else default_sub
        except ValueError:
            item_subtotal = default_sub
        if item_subtotal is None:
            print("    Subtotal is required.")
            continue

        item_notes = get_optional_input("    Notes (optional — press Enter to skip): ")

        items.append({
            "title": item_title,
            "category": category,
            "quantity": qty,
            "unit_of_measure": uom,
            "unit_price": up,
            "subtotal": item_subtotal,
            "notes": item_notes,
        })

        running_total += item_subtotal
        item_num += 1
        print(f"\n    Running total: {currency} {running_total:,.2f}")

        if not confirm("\n  Add another item?"):
            break

    if not items:
        print("\n✗ At least one line item is required.")
        input("\nPress Enter to continue...")
        return

    # ---- FULL SUMMARY ----
    igv_amount = running_total * (igv_rate / 100)
    total = running_total + igv_amount
    detraccion_amount = total * (detraccion_rate / 100) if detraccion_rate else 0

    print("\n--- Cost Summary ---")
    if project:
        print(f"  Project:     {project['project_code']} — {project['name']}")
    else:
        print(f"  Project:     (SG&A — no project)")
    if entity:
        print(f"  Entity:      {entity['legal_name']}")
    print(f"  Date:        {date_str}")
    print(f"  Title:       {title}")
    partner_name = bank_account.get("partner_companies", {}).get("name", "Unknown")
    print(f"  Bank:        {bank_account['bank_name']} {bank_account['currency']} {bank_account['account_number_last4']} ({partner_name})")
    print(f"  Cost type:   {cost_type}")
    print(f"  Currency:    {currency}")
    if payment_method:
        print(f"  Payment:     {payment_method}")

    print(f"\n  Line Items:")
    for i, item in enumerate(items, start=1):
        print(f"    {i}. {item['title']} ({item['category']}) — {currency} {item['subtotal']:,.2f}")

    print(f"\n  Subtotal:      {currency} {running_total:,.2f}")
    print(f"  IGV ({igv_rate}%):   {currency} {igv_amount:,.2f}")
    print(f"  Total:         {currency} {total:,.2f}")
    if detraccion_rate:
        print(f"  Detraccion ({detraccion_rate}%): {currency} {detraccion_amount:,.2f}")

    if not confirm("\nRegister this cost?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # ---- INSERT HEADER ----
    header_data = {
        "cost_type": cost_type,
        "bank_account_id": bank_account["id"],
        "date": date_str,
        "title": title,
        "igv_rate": igv_rate,
        "currency": currency,
    }
    if project:
        header_data["project_id"] = project["id"]
    if valuation:
        header_data["valuation_id"] = valuation["id"]
    if entity:
        header_data["entity_id"] = entity["id"]
    if quote:
        header_data["quote_id"] = quote["id"]
    if detraccion_rate:
        header_data["detraccion_rate"] = detraccion_rate
    header_data["exchange_rate"] = exchange_rate
    if comprobante_type:
        header_data["comprobante_type"] = comprobante_type
    if comprobante_number:
        header_data["comprobante_number"] = comprobante_number
    if payment_method:
        header_data["payment_method"] = payment_method
    if document_ref:
        header_data["document_ref"] = document_ref
    if due_date:
        header_data["due_date"] = due_date
    if notes:
        header_data["notes"] = notes

    try:
        response = supabase.table("costs").insert(header_data).execute()
        cost_id = response.data[0]["id"]
        print(f"\n✓ Cost header registered (ID: {cost_id[:8]}...)")
    except Exception as e:
        print(f"\n✗ Error inserting cost header: {e}")
        input("\nPress Enter to continue...")
        return

    # ---- INSERT LINE ITEMS ----
    item_records = []
    for item in items:
        record = {
            "cost_id": cost_id,
            "title": item["title"],
            "category": item["category"],
            "subtotal": item["subtotal"],
        }
        if item["quantity"]:
            record["quantity"] = item["quantity"]
        if item["unit_of_measure"]:
            record["unit_of_measure"] = item["unit_of_measure"]
        if item["unit_price"]:
            record["unit_price"] = item["unit_price"]
        if item["notes"]:
            record["notes"] = item["notes"]
        item_records.append(record)

    try:
        response = supabase.table("cost_items").insert(item_records).execute()
        print(f"✓ {len(response.data)} line items registered.")
    except Exception as e:
        print(f"\n✗ Error inserting line items: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Import Costs
# ============================================================

def _load_cost_lookups():
    """Pre-load all FK lookup tables for cost import."""
    projects = supabase.table("projects").select("id, project_code").eq("is_active", True).execute()
    entities = supabase.table("entities").select("id, document_number").eq("is_active", True).execute()
    bank_accounts = supabase.table("bank_accounts").select("id, bank_name, account_number_last4").eq("is_active", True).execute()
    valuations = supabase.table("valuations").select("id, project_id, valuation_number").execute()
    quotes = supabase.table("quotes").select("id, document_ref").execute()

    # Build project_code -> id map
    project_map = {r["project_code"]: r["id"] for r in projects.data}

    # Build composite bank account lookup: "BankName-Last4" -> id
    bank_map = {}
    for r in bank_accounts.data:
        key = f"{r['bank_name']}-{r['account_number_last4']}"
        bank_map[key] = r["id"]

    # Build composite valuation lookup: (project_id, valuation_number) -> id
    valuation_map = {}
    for r in valuations.data:
        valuation_map[(r["project_id"], r["valuation_number"])] = r["id"]

    # Build quote lookup: document_ref -> id
    quote_map = {}
    for r in quotes.data:
        if r.get("document_ref"):
            quote_map[r["document_ref"]] = r["id"]

    return {
        "projects": project_map,
        "entities": {r["document_number"]: r["id"] for r in entities.data},
        "bank_accounts": bank_map,
        "valuations": valuation_map,
        "quotes": quote_map,
    }


def _validate_cost_row(row_num, row, errors, lookups):
    """Validate a single cost row."""
    validate_required(row_num, row, "bank_name", errors)
    validate_required(row_num, row, "bank_account_last4", errors)
    validate_required(row_num, row, "cost_type", errors)
    validate_required(row_num, row, "date", errors)
    validate_required(row_num, row, "title", errors)
    validate_required(row_num, row, "igv_rate", errors)
    validate_required(row_num, row, "currency", errors)
    validate_required(row_num, row, "exchange_rate", errors)

    validate_enum(row_num, row, "cost_type", ["project_cost", "sga"], errors)
    validate_enum(row_num, row, "currency", ["USD", "PEN"], errors)
    validate_enum(row_num, row, "comprobante_type",
                  ["factura", "boleta", "recibo_por_honorarios", "liquidacion_de_compra", "planilla_jornales", "none"], errors)
    validate_enum(row_num, row, "payment_method", ["bank_transfer", "cash", "check"], errors)

    validate_lookup(row_num, row, "project_code", lookups["projects"], errors)
    validate_lookup(row_num, row, "entity_document_number", lookups["entities"], errors)

    # Composite bank account lookup
    bank_name = row.get("bank_name")
    last4 = row.get("bank_account_last4")
    if bank_name and last4 and not pd.isna(bank_name) and not pd.isna(last4):
        key = f"{str(bank_name).strip()}-{str(last4).strip()}"
        if key not in lookups["bank_accounts"]:
            errors.append((row_num, "bank_name", f"Bank account {key} not found in database"))

    # Composite valuation lookup
    proj_code = row.get("project_code")
    val_num = row.get("valuation_number")
    if proj_code and val_num and not pd.isna(proj_code) and not pd.isna(val_num):
        project_id = lookups["projects"].get(str(proj_code).strip())
        if project_id:
            try:
                vn = int(float(val_num))
                if (project_id, vn) not in lookups["valuations"]:
                    errors.append((row_num, "valuation_number", f"Valuation #{vn} not found for {proj_code}"))
            except (ValueError, TypeError):
                errors.append((row_num, "valuation_number", "Must be a number"))

    # Quote lookup
    validate_lookup(row_num, row, "quote_document_ref", lookups["quotes"], errors)

    validate_date(row_num, row, "date", errors)
    validate_date(row_num, row, "due_date", errors)
    validate_number(row_num, row, "igv_rate", errors)
    validate_number(row_num, row, "detraccion_rate", errors)
    validate_number(row_num, row, "exchange_rate", errors)


def _build_cost_record(row, lookups):
    """Convert a spreadsheet row to a database record."""
    # Required: bank account (composite)
    bank_key = f"{str(row['bank_name']).strip()}-{str(row['bank_account_last4']).strip()}"

    data = {
        "bank_account_id": lookups["bank_accounts"][bank_key],
        "cost_type": str(row["cost_type"]).strip(),
        "date": pd.Timestamp(row["date"]).strftime("%Y-%m-%d"),
        "title": str(row["title"]).strip(),
        "igv_rate": float(row["igv_rate"]),
        "currency": str(row["currency"]).strip(),
        "exchange_rate": float(row["exchange_rate"]),
    }

    # FK lookups (optional)
    proj_code = row.get("project_code")
    if proj_code and not pd.isna(proj_code) and str(proj_code).strip():
        data["project_id"] = lookups["projects"][str(proj_code).strip()]

    entity_doc = row.get("entity_document_number")
    if entity_doc and not pd.isna(entity_doc) and str(entity_doc).strip():
        data["entity_id"] = lookups["entities"][str(entity_doc).strip()]

    val_num = row.get("valuation_number")
    if proj_code and val_num and not pd.isna(proj_code) and not pd.isna(val_num):
        project_id = lookups["projects"].get(str(proj_code).strip())
        if project_id:
            vn = int(float(val_num))
            val_key = (project_id, vn)
            if val_key in lookups["valuations"]:
                data["valuation_id"] = lookups["valuations"][val_key]

    quote_ref = row.get("quote_document_ref")
    if quote_ref and not pd.isna(quote_ref) and str(quote_ref).strip():
        data["quote_id"] = lookups["quotes"][str(quote_ref).strip()]

    # Optional numeric fields
    for field in ("detraccion_rate",):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = float(val)

    # Optional string fields
    for field in ("comprobante_type", "comprobante_number", "payment_method", "document_ref", "notes"):
        val = row.get(field)
        if val is not None and not pd.isna(val) and str(val).strip():
            data[field] = str(val).strip()

    # Optional date fields
    for field in ("due_date",):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = pd.Timestamp(val).strftime("%Y-%m-%d")

    return data


def import_costs():
    """Import costs from an Excel spreadsheet."""
    clear_screen()
    print("\n=== Import Costs ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return

    print(f"Found {len(df)} data rows.")

    lookups = _load_cost_lookups()

    errors = []
    for idx, row in df.iterrows():
        excel_row = idx + DATA_START_ROW
        _validate_cost_row(excel_row, row, errors, lookups)

    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)

    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        print(f"    {i+1}. {row.get('project_code', 'SGA')} — {row.get('title', '')}")

    if not confirm(f"\nImport {len(df)} costs?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    try:
        records = [_build_cost_record(row, lookups) for _, row in df.iterrows()]
        response = supabase.table("costs").insert(records).execute()
        print(f"\n✓ {len(response.data)} costs imported successfully.")
    except Exception as e:
        print(f"\n✗ Error during import: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Import Cost Items
# ============================================================

def _load_cost_item_lookups():
    """Pre-load lookup tables for cost item import."""
    costs = supabase.table("costs").select("id, document_ref").execute()
    return {
        "costs": {r["document_ref"]: r["id"] for r in costs.data if r.get("document_ref")},
    }


def _validate_cost_item_row(row_num, row, errors, lookups):
    """Validate a single cost item row."""
    validate_required(row_num, row, "cost_document_ref", errors)
    validate_required(row_num, row, "title", errors)
    validate_required(row_num, row, "category", errors)
    validate_required(row_num, row, "subtotal", errors)

    validate_enum(row_num, row, "category", ALL_CATEGORIES, errors)
    validate_lookup(row_num, row, "cost_document_ref", lookups["costs"], errors)
    validate_number(row_num, row, "quantity", errors)
    validate_number(row_num, row, "unit_price", errors)
    validate_number(row_num, row, "subtotal", errors)


def _build_cost_item_record(row, lookups):
    """Convert a spreadsheet row to a database record."""
    data = {
        "cost_id": lookups["costs"][str(row["cost_document_ref"]).strip()],
        "title": str(row["title"]).strip(),
        "category": str(row["category"]).strip(),
        "subtotal": float(row["subtotal"]),
    }

    for field in ("quantity", "unit_price"):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = float(val)

    for field in ("unit_of_measure", "notes"):
        val = row.get(field)
        if val is not None and not pd.isna(val) and str(val).strip():
            data[field] = str(val).strip()

    return data


def import_cost_items():
    """Import cost items from an Excel spreadsheet."""
    clear_screen()
    print("\n=== Import Cost Items ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return

    print(f"Found {len(df)} data rows.")

    lookups = _load_cost_item_lookups()

    errors = []
    for idx, row in df.iterrows():
        excel_row = idx + DATA_START_ROW
        _validate_cost_item_row(excel_row, row, errors, lookups)

    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)

    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        print(f"    {i+1}. [{row.get('cost_document_ref', '')}] {row.get('title', '')} — {row.get('subtotal', '')}")

    if not confirm(f"\nImport {len(df)} cost items?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    try:
        records = [_build_cost_item_record(row, lookups) for _, row in df.iterrows()]
        response = supabase.table("cost_items").insert(records).execute()
        print(f"\n✓ {len(response.data)} cost items imported successfully.")
    except Exception as e:
        print(f"\n✗ Error during import: {e}")

    input("\nPress Enter to continue...")

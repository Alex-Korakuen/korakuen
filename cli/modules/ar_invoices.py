#!/usr/bin/env python3
"""
Module: ar_invoices.py
Purpose: All AR invoice operations — add single, import from Excel
Tables: ar_invoices
"""

import pandas as pd
from openpyxl import load_workbook

from lib.db import supabase
from lib.helpers import (
    get_input, get_optional_input, confirm, list_choices, clear_screen,
    get_currency, get_exchange_rate, select_project,
)
from lib.import_helpers import (
    DATA_START_ROW,
    clear_highlighting, apply_error_highlighting,
    validate_required, validate_enum, validate_lookup,
    validate_date, validate_number, validate_boolean,
    print_errors,
)


def menu():
    """Submenu for AR invoice operations. Called by main.py."""
    while True:
        clear_screen()
        print("\n=== AR Invoices ===\n")
        print("1. Add AR invoice")
        print("2. Import AR invoices from Excel")
        print("3. Back")

        choice = get_input("\nSelect option: ")

        if choice == "1":
            add_ar_invoice()
        elif choice == "2":
            import_ar_invoices()
        elif choice == "3":
            return


# ============================================================
# Add AR Invoice
# ============================================================

def add_ar_invoice():
    """Register a single AR invoice interactively."""
    clear_screen()
    print("\n=== Add AR Invoice ===\n")

    # --- Select project ---
    project = select_project()
    if not project:
        return

    # --- Select valuation (exclude those with existing AR invoices) ---
    all_vals = (
        supabase.table("valuations")
        .select("id, valuation_number, period_month, period_year, status")
        .eq("project_id", project["id"])
        .order("valuation_number")
        .execute()
    )
    existing_ar = (
        supabase.table("ar_invoices")
        .select("valuation_id")
        .eq("project_id", project["id"])
        .execute()
    )
    used_val_ids = {r["valuation_id"] for r in existing_ar.data}
    available_vals = [v for v in all_vals.data if v["id"] not in used_val_ids]

    if not available_vals:
        print("\n  No valuations available (all already have AR invoices).")
        input("\nPress Enter to continue...")
        return

    print("\n  Available valuations:")
    for i, v in enumerate(available_vals, start=1):
        print(f"    {i}. Val #{v['valuation_number']} — {v['period_month']}/{v['period_year']} ({v['status']})")
    print()

    val_num = get_input("  Select valuation number: ")
    try:
        valuation = available_vals[int(val_num) - 1]
    except (ValueError, IndexError):
        print("\n✗ Invalid selection.")
        input("\nPress Enter to continue...")
        return

    # --- Select bank account ---
    bank_accounts = (
        supabase.table("bank_accounts")
        .select("id, bank_name, account_number_last4, currency, is_detraccion_account, partner_companies(name)")
        .eq("is_active", True)
        .eq("is_detraccion_account", False)
        .execute()
    )
    if not bank_accounts.data:
        print("\n  No regular bank accounts found.")
        input("\nPress Enter to continue...")
        return

    print("\n  Available bank accounts (receipt):")
    for i, ba in enumerate(bank_accounts.data, start=1):
        partner = ba.get("partner_companies", {}).get("name", "Unknown")
        print(f"    {i}. {ba['bank_name']} {ba['currency']} {ba['account_number_last4']} ({partner})")
    print()

    bank_num = get_input("  Select bank account number: ")
    try:
        bank_account = bank_accounts.data[int(bank_num) - 1]
    except (ValueError, IndexError):
        print("\n✗ Invalid selection.")
        input("\nPress Enter to continue...")
        return

    # --- Select client entity ---
    from modules.entities import _search_and_select_entity
    print("\n  Client entity:")
    entity = _search_and_select_entity()
    if not entity:
        return

    # --- Select partner company ---
    partners = (
        supabase.table("partner_companies")
        .select("id, name, ruc")
        .eq("is_active", True)
        .execute()
    )
    list_choices("Partner companies", partners.data, display=["name", "ruc"])
    partner_num = get_input("  Select partner company number: ")
    try:
        partner = partners.data[int(partner_num) - 1]
    except (ValueError, IndexError):
        print("\n✗ Invalid selection.")
        input("\nPress Enter to continue...")
        return

    # --- Invoice details ---
    invoice_number = get_input("\n  Invoice number: ")

    print("\n  Comprobante types: factura, boleta, recibo_por_honorarios")
    comprobante_type = get_input("  Comprobante type (default: factura): ").lower() or "factura"
    while comprobante_type not in ("factura", "boleta", "recibo_por_honorarios"):
        print("  Must be factura, boleta, or recibo_por_honorarios.")
        comprobante_type = get_input("  Comprobante type: ").lower()

    invoice_date = get_input("  Invoice date (YYYY-MM-DD): ")
    due_date = get_optional_input("  Due date (YYYY-MM-DD, optional — press Enter to skip): ")

    # --- Financial ---
    subtotal_input = get_input("\n  Subtotal: ")
    try:
        subtotal = float(subtotal_input)
    except ValueError:
        print("  Invalid number.")
        input("\nPress Enter to continue...")
        return

    igv_input = get_input("  IGV rate % (default: 18): ") or "18"
    try:
        igv_rate = float(igv_input)
    except ValueError:
        igv_rate = 18.0

    detraccion_input = get_optional_input("  Detraccion rate % (optional — press Enter to skip): ")
    detraccion_rate = None
    if detraccion_input:
        try:
            detraccion_rate = float(detraccion_input)
        except ValueError:
            detraccion_rate = None

    # --- Retencion ---
    retencion_applicable = confirm("  Retencion applicable?")
    retencion_rate = None
    if retencion_applicable:
        ret_input = get_input("  Retencion rate % (default: 3): ") or "3"
        try:
            retencion_rate = float(ret_input)
        except ValueError:
            retencion_rate = 3.0

    # --- Currency ---
    print("\n  Currencies: USD, PEN")
    currency = get_currency()
    exchange_rate = get_exchange_rate()

    document_ref = get_optional_input("  Document ref (e.g. PRY001-AR-001, optional — press Enter to skip): ")
    is_internal_settlement = confirm("  Is internal settlement (partner-to-partner)?")
    notes = get_optional_input("  Notes (optional — press Enter to skip): ")

    # --- Tax breakdown (display-only) ---
    igv_amount = subtotal * (igv_rate / 100)
    gross_total = subtotal + igv_amount
    detraccion_amount = gross_total * (detraccion_rate / 100) if detraccion_rate else 0
    retencion_amount = gross_total * (retencion_rate / 100) if retencion_rate else 0
    net_receivable = gross_total - detraccion_amount - retencion_amount

    print("\n--- Tax Breakdown ---")
    print(f"  Subtotal:        {currency} {subtotal:>12,.2f}")
    print(f"  IGV ({igv_rate}%):      {currency} {igv_amount:>12,.2f}")
    print(f"  Gross Total:     {currency} {gross_total:>12,.2f}")
    if detraccion_rate:
        print(f"  Detraccion ({detraccion_rate}%): {currency} {detraccion_amount:>12,.2f}")
    if retencion_rate:
        print(f"  Retencion ({retencion_rate}%):  {currency} {retencion_amount:>12,.2f}")
    print(f"  Net Receivable:  {currency} {net_receivable:>12,.2f}")

    # --- Full Summary ---
    print("\n--- Summary ---")
    print(f"  Project:    {project['project_code']} — {project['name']}")
    print(f"  Valuation:  #{valuation['valuation_number']}")
    print(f"  Client:     {entity['legal_name']}")
    print(f"  Partner:    {partner['name']}")
    print(f"  Invoice:    {invoice_number} ({comprobante_type})")
    print(f"  Date:       {invoice_date}")
    if is_internal_settlement:
        print(f"  Internal:   Yes (partner settlement)")

    if not confirm("\nRegister this AR invoice?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    # --- Insert ---
    data = {
        "project_id": project["id"],
        "valuation_id": valuation["id"],
        "bank_account_id": bank_account["id"],
        "entity_id": entity["id"],
        "partner_company_id": partner["id"],
        "invoice_number": invoice_number,
        "comprobante_type": comprobante_type,
        "invoice_date": invoice_date,
        "subtotal": subtotal,
        "igv_rate": igv_rate,
        "retencion_applicable": retencion_applicable,
        "currency": currency,
        "is_internal_settlement": is_internal_settlement,
        "retencion_verified": False,
    }
    if due_date:
        data["due_date"] = due_date
    if detraccion_rate:
        data["detraccion_rate"] = detraccion_rate
    if retencion_rate:
        data["retencion_rate"] = retencion_rate
    data["exchange_rate"] = exchange_rate
    if document_ref:
        data["document_ref"] = document_ref
    if notes:
        data["notes"] = notes

    try:
        response = supabase.table("ar_invoices").insert(data).execute()
        print(f"\n✓ AR invoice registered (ID: {response.data[0]['id'][:8]}...)")
    except Exception as e:
        print(f"\n✗ Error: {e}")

    input("\nPress Enter to continue...")


# ============================================================
# Import AR Invoices
# ============================================================

def _load_ar_lookups():
    """Pre-load all FK lookup tables for AR invoice import."""
    projects = supabase.table("projects").select("id, project_code").eq("is_active", True).execute()
    entities = supabase.table("entities").select("id, document_number").eq("is_active", True).execute()
    bank_accounts = supabase.table("bank_accounts").select("id, bank_name, account_number_last4").eq("is_active", True).execute()
    valuations = supabase.table("valuations").select("id, project_id, valuation_number").execute()
    partners = supabase.table("partner_companies").select("id, name").eq("is_active", True).execute()

    project_map = {r["project_code"]: r["id"] for r in projects.data}

    bank_map = {}
    for r in bank_accounts.data:
        key = f"{r['bank_name']}-{r['account_number_last4']}"
        bank_map[key] = r["id"]

    valuation_map = {}
    for r in valuations.data:
        valuation_map[(r["project_id"], r["valuation_number"])] = r["id"]

    return {
        "projects": project_map,
        "entities": {r["document_number"]: r["id"] for r in entities.data},
        "bank_accounts": bank_map,
        "valuations": valuation_map,
        "partners": {r["name"]: r["id"] for r in partners.data},
    }


def _parse_bool(val):
    """Parse a boolean value from Excel (handles Python bool, string, etc.)."""
    if isinstance(val, bool):
        return val
    if pd.isna(val):
        return False
    return str(val).strip().lower() == "true"


def _validate_ar_row(row_num, row, errors, lookups):
    """Validate a single AR invoice row."""
    validate_required(row_num, row, "project_code", errors)
    validate_required(row_num, row, "valuation_number", errors)
    validate_required(row_num, row, "bank_name", errors)
    validate_required(row_num, row, "bank_account_last4", errors)
    validate_required(row_num, row, "entity_document_number", errors)
    validate_required(row_num, row, "partner_company_name", errors)
    validate_required(row_num, row, "invoice_number", errors)
    validate_required(row_num, row, "comprobante_type", errors)
    validate_required(row_num, row, "invoice_date", errors)
    validate_required(row_num, row, "subtotal", errors)
    validate_required(row_num, row, "igv_rate", errors)
    validate_required(row_num, row, "currency", errors)
    validate_required(row_num, row, "exchange_rate", errors)
    validate_required(row_num, row, "retencion_applicable", errors)
    validate_required(row_num, row, "is_internal_settlement", errors)
    validate_required(row_num, row, "retencion_verified", errors)

    validate_enum(row_num, row, "comprobante_type", ["factura", "boleta", "recibo_por_honorarios"], errors)
    validate_enum(row_num, row, "currency", ["USD", "PEN"], errors)

    validate_lookup(row_num, row, "project_code", lookups["projects"], errors)
    validate_lookup(row_num, row, "entity_document_number", lookups["entities"], errors)
    validate_lookup(row_num, row, "partner_company_name", lookups["partners"], errors)

    # Composite bank account lookup
    bank_name = row.get("bank_name")
    last4 = row.get("bank_account_last4")
    if bank_name and last4 and not pd.isna(bank_name) and not pd.isna(last4):
        key = f"{str(bank_name).strip()}-{str(last4).strip()}"
        if key not in lookups["bank_accounts"]:
            errors.append((row_num, "bank_name", f"Bank account {key} not found"))

    # Composite valuation lookup
    proj_code = row.get("project_code")
    val_num = row.get("valuation_number")
    if proj_code and val_num and not pd.isna(proj_code) and not pd.isna(val_num):
        project_id = lookups["projects"].get(str(proj_code).strip())
        if project_id:
            try:
                vn = int(float(val_num))
                if (project_id, vn) not in lookups["valuations"]:
                    errors.append((row_num, "valuation_number", f"Valuation #{vn} not found for {proj_code}"))
            except (ValueError, TypeError):
                errors.append((row_num, "valuation_number", "Must be a number"))

    validate_date(row_num, row, "invoice_date", errors)
    validate_date(row_num, row, "due_date", errors)
    validate_number(row_num, row, "subtotal", errors)
    validate_number(row_num, row, "igv_rate", errors)
    validate_number(row_num, row, "detraccion_rate", errors)
    validate_number(row_num, row, "retencion_rate", errors)
    validate_number(row_num, row, "exchange_rate", errors)

    validate_boolean(row_num, row, "retencion_applicable", errors)
    validate_boolean(row_num, row, "is_internal_settlement", errors)
    validate_boolean(row_num, row, "retencion_verified", errors)

    # Cross-field: if retencion_applicable, retencion_rate must be present
    ret_app = row.get("retencion_applicable")
    if ret_app and not pd.isna(ret_app) and _parse_bool(ret_app):
        ret_rate = row.get("retencion_rate")
        if ret_rate is None or pd.isna(ret_rate):
            errors.append((row_num, "retencion_rate", "Required when retencion_applicable is true"))


def _build_ar_record(row, lookups):
    """Convert a spreadsheet row to a database record."""
    proj_code = str(row["project_code"]).strip()
    project_id = lookups["projects"][proj_code]

    bank_key = f"{str(row['bank_name']).strip()}-{str(row['bank_account_last4']).strip()}"
    vn = int(float(row["valuation_number"]))

    data = {
        "project_id": project_id,
        "valuation_id": lookups["valuations"][(project_id, vn)],
        "bank_account_id": lookups["bank_accounts"][bank_key],
        "entity_id": lookups["entities"][str(row["entity_document_number"]).strip()],
        "partner_company_id": lookups["partners"][str(row["partner_company_name"]).strip()],
        "invoice_number": str(row["invoice_number"]).strip(),
        "comprobante_type": str(row["comprobante_type"]).strip(),
        "invoice_date": pd.Timestamp(row["invoice_date"]).strftime("%Y-%m-%d"),
        "subtotal": float(row["subtotal"]),
        "igv_rate": float(row["igv_rate"]),
        "retencion_applicable": _parse_bool(row.get("retencion_applicable")),
        "currency": str(row["currency"]).strip(),
        "exchange_rate": float(row["exchange_rate"]),
        "is_internal_settlement": _parse_bool(row.get("is_internal_settlement")),
        "retencion_verified": _parse_bool(row.get("retencion_verified")),
    }

    # Optional numeric
    for field in ("detraccion_rate", "retencion_rate"):
        val = row.get(field)
        if val is not None and not pd.isna(val):
            data[field] = float(val)

    # Optional string
    for field in ("document_ref", "notes"):
        val = row.get(field)
        if val is not None and not pd.isna(val) and str(val).strip():
            data[field] = str(val).strip()

    # Optional date
    due = row.get("due_date")
    if due is not None and not pd.isna(due):
        data["due_date"] = pd.Timestamp(due).strftime("%Y-%m-%d")

    return data


def import_ar_invoices():
    """Import AR invoices from an Excel spreadsheet."""
    clear_screen()
    print("\n=== Import AR Invoices ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return

    print(f"Found {len(df)} data rows.")

    lookups = _load_ar_lookups()

    errors = []
    for idx, row in df.iterrows():
        excel_row = idx + DATA_START_ROW
        _validate_ar_row(excel_row, row, errors, lookups)

    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)

    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        print(f"    {i+1}. {row.get('project_code', '')} — {row.get('invoice_number', '')} — {row.get('subtotal', '')}")

    if not confirm(f"\nImport {len(df)} AR invoices?"):
        print("Cancelled.")
        input("\nPress Enter to continue...")
        return

    try:
        records = [_build_ar_record(row, lookups) for _, row in df.iterrows()]
        response = supabase.table("ar_invoices").insert(records).execute()
        print(f"\n✓ {len(response.data)} AR invoices imported successfully.")
    except Exception as e:
        print(f"\n✗ Error during import: {e}")

    input("\nPress Enter to continue...")

"""
Shared import validation and error highlighting utilities.
Used by all CLI modules that have Excel import functions.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from lib.db import supabase
from lib.helpers import get_input

# Error highlighting — dark red fill for invalid cells
RED_FILL = PatternFill(fill_type="solid", fgColor="8B0000")
NO_FILL = PatternFill(fill_type=None)

# Excel template layout: rows 1-4 are headers, data starts at row 5
DATA_START_ROW = 5


def is_empty(val):
    """Check if a cell value is empty/NaN."""
    return val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == ""


def cell_str(val):
    """Normalize a cell value to a stripped string, or empty string if empty."""
    if is_empty(val):
        return ""
    return str(val).strip()


def opt_str(row, field):
    """Extract optional string field from Excel row. Returns str or None."""
    val = row.get(field)
    if is_empty(val):
        return None
    return str(val).strip() or None


def opt_float(row, field):
    """Extract optional numeric field from Excel row. Returns float or None."""
    val = row.get(field)
    if is_empty(val):
        return None
    return float(val)


def opt_date(row, field):
    """Extract optional date field from Excel row. Returns YYYY-MM-DD str or None."""
    val = row.get(field)
    if is_empty(val):
        return None
    return pd.Timestamp(val).strftime("%Y-%m-%d")


def load_excel_file(title):
    """Prompt for Excel file path, read it, and return (df, file_path).

    Handles the common import preamble: clear screen, print title, prompt for
    file path, read with pd.read_excel (skipping template header rows 2-4),
    check for empty data, and print row count.

    Returns (df, file_path) on success, or None on failure.
    """
    from lib.helpers import clear_screen
    clear_screen()
    print(f"\n=== {title} ===\n")

    file_path = get_input("Enter path to Excel file (or drag file into terminal): ").strip().strip("'\"")

    try:
        df = pd.read_excel(file_path, header=0, skiprows=[1, 2, 3], engine="openpyxl")
    except Exception as e:
        print(f"\n✗ Error reading file: {e}")
        input("\nPress Enter to continue...")
        return None

    if df.empty:
        print("✗ No data rows found in file.")
        input("\nPress Enter to continue...")
        return None

    print(f"Found {len(df)} data rows.")
    return df, file_path


def print_import_summary(file_path, df, format_row):
    """Print the standard import summary header and first 3 preview rows.

    Args:
        file_path: Path to the Excel file being imported.
        df: The pandas DataFrame of import data.
        format_row: Callable(i, row) -> str that formats one preview row.
            i is 1-indexed, row is a pandas Series.
    """
    print(f"\n--- Summary ---")
    print(f"  File:    {file_path}")
    print(f"  Records: {len(df)}")
    print(f"\n  First 3 rows:")
    for i, (_, row) in enumerate(df.head(3).iterrows()):
        print(f"    {format_row(i + 1, row)}")


def clear_highlighting(ws):
    """Reset all cell fills from DATA_START_ROW onward."""
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        for cell in row:
            cell.fill = NO_FILL


def apply_error_highlighting(ws, errors, headers):
    """Apply RED_FILL to cells that failed validation.

    Args:
        ws: openpyxl worksheet.
        errors: List of (row_num, column_name, message) tuples.
        headers: List of column names from row 1.
    """
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}
    for row_num, col_name, _message in errors:
        col_idx = col_map.get(col_name)
        if col_idx:
            ws.cell(row=row_num, column=col_idx).fill = RED_FILL


def validate_required(row_num, row, field, errors):
    """Check that a required field is not empty or NaN."""
    val = row.get(field)
    if is_empty(val):
        errors.append((row_num, field, "Required field is empty"))


def validate_enum(row_num, row, field, allowed_values, errors):
    """Check that a field value is in the allowed list. Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    if str(val).strip() not in allowed_values:
        errors.append((row_num, field, f"Must be one of: {', '.join(allowed_values)}"))


def validate_lookup(row_num, row, field, lookup_dict, errors):
    """Check that a field value exists in the lookup dictionary. Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    if str(val).strip() not in lookup_dict:
        errors.append((row_num, field, "Not found in database"))


def validate_date(row_num, row, field, errors):
    """Check that a field value is a valid date. Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    try:
        pd.Timestamp(val)
    except (ValueError, TypeError):
        errors.append((row_num, field, "Invalid date format"))


def validate_number(row_num, row, field, errors):
    """Check that a field value is numeric. Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    try:
        float(val)
    except (ValueError, TypeError):
        errors.append((row_num, field, "Must be a number"))


def validate_nonneg_number(row_num, row, field, errors):
    """Check that a field value is a non-negative number. Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    try:
        num = float(val)
        if num < 0:
            errors.append((row_num, field, "Must not be negative"))
    except (ValueError, TypeError):
        errors.append((row_num, field, "Must be a number"))


def validate_exchange_rate(row_num, row, field, errors):
    """Check that exchange rate is within a reasonable range. Skip if empty."""
    from lib.helpers import EXCHANGE_RATE_MIN, EXCHANGE_RATE_MAX
    val = row.get(field)
    if is_empty(val):
        return
    try:
        rate = float(val)
        if rate > 0 and not (EXCHANGE_RATE_MIN <= rate <= EXCHANGE_RATE_MAX):
            errors.append((row_num, field, f"Rate {rate} is outside typical range ({EXCHANGE_RATE_MIN}–{EXCHANGE_RATE_MAX}) — verify value"))
    except (ValueError, TypeError):
        pass  # already caught by validate_nonneg_number


def parse_bool(val):
    """Parse a boolean value from Excel (handles Python bool, string, NaN).
    Returns False for empty/NaN values."""
    if isinstance(val, bool):
        return val
    if pd.isna(val):
        return False
    return str(val).strip().lower() == "true"


def validate_boolean(row_num, row, field, errors):
    """Check that a field value is a boolean (true/false). Skip if empty."""
    val = row.get(field)
    if is_empty(val):
        return
    # Pandas may read Excel booleans as Python bool
    if isinstance(val, bool):
        return
    if str(val).strip().lower() not in ("true", "false"):
        errors.append((row_num, field, "Must be true or false"))


def validate_bank_account(row_num, row, lookups, errors):
    """Check that bank_account label matches a known bank account. Skip if empty."""
    val = row.get("bank_account")
    if not is_empty(val):
        label = cell_str(val)
        if label not in lookups["bank_accounts"]:
            errors.append((row_num, "bank_account", f"Bank account '{label}' not found"))


def validate_partner_company(row_num, row, lookups, errors):
    """Check that partner_company name matches a known partner. Skip if empty."""
    val = row.get("partner_company")
    if not is_empty(val):
        name = cell_str(val)
        if name not in lookups["partners"]:
            errors.append((row_num, "partner_company", f"Partner company '{name}' not found"))


def print_errors(errors, file_path):
    """Print a formatted error table to the terminal."""
    print(f"\n✗ {len(errors)} validation error(s) found:\n")
    print(f"  {'Row':<8}{'Column':<30}{'Error'}")
    print(f"  {'---':<8}{'------':<30}{'-----'}")
    for row_num, col_name, message in errors:
        print(f"  {row_num:<8}{col_name:<30}{message}")
    print(f"\n  Errors highlighted in red in: {file_path}")
    print("  Fix the highlighted cells and re-run.")


def process_import_errors(file_path, errors):
    """Handle the error/clean-highlighting cycle after import validation.

    Returns True if errors were found (caller should return early),
    False if validation passed (caller should continue with import).
    """
    if errors:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        clear_highlighting(ws)
        apply_error_highlighting(ws, errors, headers)
        wb.save(file_path)
        print_errors(errors, file_path)
        input("\nPress Enter to continue...")
        return True

    # Clear previous highlighting on clean run
    wb = load_workbook(file_path)
    ws = wb.active
    clear_highlighting(ws)
    wb.save(file_path)
    return False


# ============================================================
# Shared Lookup Builders — used by import functions across modules
# ============================================================

def load_project_map():
    """Return {project_code: id} for all active projects."""
    result = supabase.table("projects").select("id, project_code").eq("is_active", True).execute()
    return {r["project_code"]: r["id"] for r in result.data}


def load_entity_map():
    """Return {document_number: id} for all active entities."""
    result = supabase.table("entities").select("id, document_number").eq("is_active", True).execute()
    return {r["document_number"]: r["id"] for r in result.data}


def load_bank_account_map():
    """Return {label: id} for all active bank accounts."""
    result = supabase.table("bank_accounts").select("id, label").eq("is_active", True).execute()
    return {r["label"]: r["id"] for r in result.data}


def load_partner_map():
    """Return {name: id} for all active partner companies."""
    result = supabase.table("partner_companies").select("id, name").eq("is_active", True).execute()
    return {r["name"]: r["id"] for r in result.data}


def load_quote_map():
    """Return {document_ref: id} for all quotes with a document_ref."""
    result = supabase.table("quotes").select("id, document_ref").execute()
    return {r["document_ref"]: r["id"] for r in result.data if r.get("document_ref")}
